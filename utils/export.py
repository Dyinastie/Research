from io import BytesIO

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    Alignment,
    Border,
    Side
)


def export_excel(df):

    # =====================================
    # Workbook
    # =====================================
    wb = Workbook()

    ws = wb.active

    ws.title = "Hasil Prediksi"

    # =====================================
    # Border
    # =====================================
    thin = Side(
        style="thin",
        color="000000"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # =====================================
    # Header Atas
    # =====================================
    ws["A2"] = (
        "FORM IDENTIFIKASI KAMUS "
        "USULAN MUSRENBANG"
    )

    ws["A3"] = (
        "DALAM RANGKA PENYUSUNAN "
        "KAMUS USULAN MUSRENBANG"
    )

    # =====================================
    # Merge Header
    # =====================================
    ws.merge_cells("A2:L2")
    ws.merge_cells("A3:L3")

    # =====================================
    # Style Header Atas
    # =====================================
    for cell in ["A2", "A3"]:

        ws[cell].font = Font(
            name="Arial",
            size=14,
            bold=True
        )

        ws[cell].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # =====================================
    # Header Tabel
    # =====================================
    headers = [
        "Arah Kebijakan Pembangunan Kota Malang",
        "No.",
        "Kode",
        "Kecamatan",
        "Kelurahan",
        "Permasalahan",
        "Penyebab",
        "Lokasi",
        "Usulan Kamus",
        "Keterangan",
        "OPD_1",
        "OPD_2"
    ]

    # =====================================
    # Tulis Header Tabel
    # =====================================
    header_row = 5

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(
            row=header_row,
            column=col_num
        )

        cell.value = header

        cell.font = Font(
            name="Arial",
            size=12,
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = border

    # =====================================
    # Isi Data
    # =====================================
    start_row = 6

    nomor = 1

    for idx, row in df.iterrows():

        data = [
            row.get("Arah_Kebijakan_Export",""),
            nomor,
            row.get("Kode", ""),
            row.get("Kecamatan", ""),
            row.get("Kelurahan", ""),
            row.get("Permasalahan", ""),
            row.get("Penyebab", ""),
            row.get("Lokasi", ""),
            row.get("Usulan Kamus", ""),
            row.get("Keterangan", ""),
            row.get("OPD_1", ""),
            row.get("OPD_2", "")
        ]

        for col_num, value in enumerate(data, 1):

            cell = ws.cell(
                row=start_row + idx,
                column=col_num
            )

            cell.value = value

            cell.font = Font(
                name="Arial",
                size=12
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

            cell.border = border

        nomor += 1

    # =====================================
    # Auto Width
    # =====================================
    for column_cells in ws.columns:

        try:

            length = max(
                len(str(cell.value))
                if cell.value else 0
                for cell in column_cells
            )

            ws.column_dimensions[
                column_cells[0].column_letter
            ].width = length + 5

        except:
            pass

    # =====================================
    # Row Height Header
    # =====================================
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 25

    # =====================================
    # Save
    # =====================================
    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output